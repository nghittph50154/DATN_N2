import openpyxl
from collections import defaultdict

# Precinct -> Borough mapping (zero-padded 3 digits to match Excel format)
PRECINCT_BOROUGH = {}
for p in [1,5,6,7,9,10,13,14,17,18,19,20,22,23,24,25,26,28,30,32,33,34]:
    PRECINCT_BOROUGH[f'{p:03d}'] = 'Manhattan'
for p in [40,41,42,43,44,45,46,47,48,49,50,52]:
    PRECINCT_BOROUGH[f'{p:03d}'] = 'The Bronx'
for p in [60,61,62,63,66,67,68,69,70,71,72,73,75,76,77,78,79,81,83,84,88,90,94]:
    PRECINCT_BOROUGH[f'{p:03d}'] = 'Brooklyn'
for p in [100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116]:
    PRECINCT_BOROUGH[f'{p:03d}'] = 'Queens'
for p in [120,121,122,123]:
    PRECINCT_BOROUGH[f'{p:03d}'] = 'Staten Island'

quarters = [
    'cs-report-q1-2025.xlsx',
    'cs-report-q2-2025.xlsx',
    'cs-report-q3-2025.xlsx',
    'cs-report-q4-2025.xlsx',
]

totals = defaultdict(lambda: {'felony': 0, 'misdemeanor': 0, 'total': 0})

for qfile in quarters:
    wb = openpyxl.load_workbook(f'D:\\{qfile}', data_only=True)
    sheet_name = next((s for s in wb.sheetnames if 'Crime Complaints' in s), None)
    if not sheet_name:
        print(f'WARN: No Crime Complaints in {qfile}')
        continue
    ws = wb[sheet_name]
    q_total = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        pct = str(row[0]).strip() if row[0] else ''
        if pct in PRECINCT_BOROUGH:
            borough = PRECINCT_BOROUGH[pct]
            try:
                felony = int(row[1]) if row[1] else 0
                misdem = int(row[4]) if row[4] else 0
                total  = int(row[7]) if row[7] else 0
                totals[borough]['felony']      += felony
                totals[borough]['misdemeanor'] += misdem
                totals[borough]['total']       += total
                q_total += total
            except:
                pass
    print(f'  Parsed {qfile}: total complaints = {q_total:,}')

print('\nFull Year 2025 Crime Data by Borough:')
print(f'{"Borough":<15} {"Felony":>10} {"Misdemeanor":>13} {"Total":>10}')
print('-' * 52)
for boro in ['Manhattan', 'Brooklyn', 'Queens', 'The Bronx', 'Staten Island']:
    d = totals[boro]
    print(f'{boro:<15} {d["felony"]:>10,} {d["misdemeanor"]:>13,} {d["total"]:>10,}')
